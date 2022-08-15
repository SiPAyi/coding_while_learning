import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('attendence.xlsx')
sheet = wb['Sheet1']

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=3, max_col=3)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "e3")

for i in range(2, 5):
    if sheet.cell(i, 3).value > 40:
        sheet.cell(i, 4).value = 'eligible'
    else:
        sheet.cell(i, 4).value = 'not eligible'

wb.save('attendence1.xlsx')
